"""
validate_llm_labels.py — Compare LLM labels against human gold set.

Uses the v3 kappa overlap sheet (100 sentences, both annotators agreed) as
the human reference. Joins to labeled_sentences.csv on sentence_id.

Computes per-label precision, recall, F1, and accuracy.
Saves results to outputs/tables/llm_validation.csv

Usage:
    python src/annotation/validate_llm_labels.py
    python src/annotation/validate_llm_labels.py \
        --gold data/annotation/kappa_overlap_person_b_v3.xlsx \
        --llm  data/annotation/labeled_sentences.csv
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from sklearn.metrics import precision_score, recall_score, f1_score

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from src.config import DATA_ANNOTATION, OUTPUTS_TABLES

GOLD_DEFAULT = DATA_ANNOTATION / "kappa_overlap_person_b_v3.xlsx"
LLM_DEFAULT  = DATA_ANNOTATION / "labeled_sentences.csv"
OUT_CSV      = OUTPUTS_TABLES  / "llm_validation.csv"

FRAME_COLS = {
    "Innovation/Progress":   "Innovation_Progress",
    "Economic Benefit":      "Economic_Benefit",
    "Risk/Harm":             "Risk_Harm",
    "Regulation/Governance": "Regulation_Governance",
    "Existential/AGI":       "Existential_AGI",
    "None":                  "None",
}


def _to_binary(v) -> int:
    if v is None:
        return 0
    s = str(v).strip()
    if s in ("", "None", "nan", "0"):
        return 0
    try:
        return 1 if float(s) >= 0.5 else 0
    except (ValueError, TypeError):
        return 0


def load_gold(path: Path) -> pd.DataFrame:
    """Read xlsx gold sheet via openpyxl; return DataFrame with sentence_id + frame cols."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rows.append(row)
    df = pd.DataFrame(rows)
    # Binarise frame columns
    for col in FRAME_COLS.values():
        if col in df.columns:
            df[col] = df[col].apply(_to_binary)
    return df[["sentence_id"] + [c for c in FRAME_COLS.values() if c in df.columns]]


def parse_args():
    p = argparse.ArgumentParser(description="Validate LLM labels against human gold set.")
    p.add_argument("--gold", default=str(GOLD_DEFAULT))
    p.add_argument("--llm",  default=str(LLM_DEFAULT))
    return p.parse_args()


def main():
    args = parse_args()

    print(f"Loading gold set from {args.gold} …")
    gold = load_gold(Path(args.gold))
    print(f"  {len(gold)} sentences")

    print(f"Loading LLM labels from {args.llm} …")
    llm_all = pd.read_csv(args.llm, dtype=str)
    for col in FRAME_COLS.values():
        if col in llm_all.columns:
            llm_all[col] = pd.to_numeric(llm_all[col], errors="coerce").fillna(0).astype(int)

    # Join on sentence_id
    merged = gold.merge(
        llm_all[["sentence_id"] + [c for c in FRAME_COLS.values() if c in llm_all.columns]],
        on="sentence_id",
        suffixes=("_human", "_llm"),
    )
    print(f"  Matched {len(merged)} / {len(gold)} gold sentences")

    if len(merged) == 0:
        print("ERROR: no sentences matched — check sentence_id alignment")
        sys.exit(1)

    # ── Per-label metrics ──────────────────────────────────────────────────────
    OUTPUTS_TABLES.mkdir(parents=True, exist_ok=True)
    results = []

    print(f"\n{'Label':<25} {'Precision':>9} {'Recall':>7} {'F1':>7} {'Accuracy':>9} "
          f"{'Human+':>7} {'LLM+':>6}")
    print("─" * 75)

    for frame, col in FRAME_COLS.items():
        human_col = col + "_human" if col + "_human" in merged.columns else col
        llm_col   = col + "_llm"   if col + "_llm"   in merged.columns else col

        if human_col not in merged.columns or llm_col not in merged.columns:
            continue

        y_true = merged[human_col].astype(int).values
        y_pred = merged[llm_col].astype(int).values

        n = len(y_true)
        accuracy  = (y_true == y_pred).sum() / n
        precision = precision_score(y_true, y_pred, zero_division=0)
        recall    = recall_score(y_true, y_pred, zero_division=0)
        f1        = f1_score(y_true, y_pred, zero_division=0)
        human_pos = int(y_true.sum())
        llm_pos   = int(y_pred.sum())

        print(f"{frame:<25} {precision:>9.3f} {recall:>7.3f} {f1:>7.3f} {accuracy:>9.3f} "
              f"{human_pos:>7} {llm_pos:>6}")

        results.append({
            "label":     frame,
            "precision": round(precision, 4),
            "recall":    round(recall, 4),
            "f1":        round(f1, 4),
            "accuracy":  round(accuracy, 4),
            "human_positive": human_pos,
            "llm_positive":   llm_pos,
            "n":              n,
        })

    # ── Macro averages ─────────────────────────────────────────────────────────
    macro_p  = sum(r["precision"] for r in results) / len(results)
    macro_r  = sum(r["recall"]    for r in results) / len(results)
    macro_f1 = sum(r["f1"]        for r in results) / len(results)
    macro_ac = sum(r["accuracy"]  for r in results) / len(results)
    print("─" * 75)
    print(f"{'MACRO AVG':<25} {macro_p:>9.3f} {macro_r:>7.3f} {macro_f1:>7.3f} {macro_ac:>9.3f}")

    # ── Disagreement inspection ────────────────────────────────────────────────
    print(f"\n── Disagreements per label ──")
    for frame, col in FRAME_COLS.items():
        human_col = col + "_human" if col + "_human" in merged.columns else col
        llm_col   = col + "_llm"   if col + "_llm"   in merged.columns else col
        if human_col not in merged.columns or llm_col not in merged.columns:
            continue
        disagree = merged[merged[human_col] != merged[llm_col]]
        fn = ((merged[human_col] == 1) & (merged[llm_col] == 0)).sum()
        fp = ((merged[human_col] == 0) & (merged[llm_col] == 1)).sum()
        if fn + fp > 0:
            print(f"  {frame:<25}  FN (missed): {fn}  FP (false alarm): {fp}")

    # ── Save ──────────────────────────────────────────────────────────────────
    out_df = pd.DataFrame(results)
    out_df.to_csv(OUT_CSV, index=False)
    print(f"\nSaved → {OUT_CSV}")

    # ── Pass/fail verdict ─────────────────────────────────────────────────────
    target = 0.80
    verdict = "PASS ✓" if macro_f1 >= target else f"FAIL (target {target:.2f})"
    print(f"\nMacro F1 = {macro_f1:.3f}  →  {verdict}")


if __name__ == "__main__":
    main()
