"""
validate_alternative_llm.py — Run an alternative LLM on the gold set and compare
to Haiku's labels. Tests robustness: if both models agree on direction and magnitude
of precision/recall trade-off, conservative recall is a general LLM property rather
than a Haiku-specific artefact.

Supported providers (auto-detected from --model prefix):
  openai  → gpt-4o-mini, gpt-4o, etc.          needs OPENAI_API_KEY
  gemini  → gemini-2.0-flash, gemini-1.5-flash  needs GEMINI_API_KEY
  groq    → llama-3.3-70b-versatile, etc.       needs GROQ_API_KEY

Usage:
    python src/annotation/validate_alternative_llm.py --model gpt-4o-mini
    python src/annotation/validate_alternative_llm.py --model gemini-2.0-flash
    python src/annotation/validate_alternative_llm.py --model llama-3.3-70b-versatile
    python src/annotation/validate_alternative_llm.py --model gpt-4o-mini --limit 20

Output: outputs/tables/llm_validation_<model>.csv
        Prints side-by-side comparison with Haiku baseline.
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from sklearn.metrics import f1_score, precision_score, recall_score

load_dotenv()

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from src.config import DATA_ANNOTATION, OUTPUTS_TABLES

GOLD_PATH = DATA_ANNOTATION / "kappa_overlap_person_a.xlsx"
HAIKU_CSV = OUTPUTS_TABLES  / "llm_validation.csv"

FRAMES = [
    "Innovation/Progress",
    "Economic Benefit",
    "Risk/Harm",
    "Regulation/Governance",
    "Existential/AGI",
    "None",
]

FRAME_COLS = {
    "Innovation/Progress":   "Innovation_Progress",
    "Economic Benefit":      "Economic_Benefit",
    "Risk/Harm":             "Risk_Harm",
    "Regulation/Governance": "Regulation_Governance",
    "Existential/AGI":       "Existential_AGI",
    "None":                  "None",
}

SYSTEM_PROMPT = """\
You are a research assistant for a computational linguistics project studying how AI
industry actors frame artificial intelligence across different contexts.

You will receive a JSON array of sentences. For each sentence, assign ONE OR MORE of
these frames, or "None" if no frame applies:

1. Innovation/Progress   — AI as transformative; EXPLICIT claim about societal benefit
                           or meaningful scientific advance. Descriptive statements
                           about how AI works are NOT enough.
2. Economic Benefit      — EXPLICIT claim about jobs, revenue, productivity, or
                           national competitiveness. Product announcements do NOT qualify.
3. Risk/Harm             — CONCRETE near-term harms: bias, job loss, misuse,
                           surveillance. Vague concern does NOT qualify.
4. Regulation/Governance — EXPLICIT institutional mechanism: law, policy framework,
                           oversight body, compliance requirement, standards body.
5. Existential/AGI       — Long-term civilizational risk, AGI, superintelligence,
                           x-risk. Near-term risks go in Risk/Harm.
None                     — Procedural, descriptive, transitional, or administrative
                           sentences. When in doubt and the frame must be INFERRED
                           rather than stated, use None.

Return ONLY a valid JSON array, one object per sentence:
  [{"id": "<sentence_id>", "labels": ["Innovation/Progress"]}, ...]
For None: {"id": "<sentence_id>", "labels": ["None"]}
No preamble, no explanation, no markdown fences. Compact JSON only — no newlines or extra whitespace."""

BATCH_SIZE       = 15
INTER_CALL_DELAY = 0.5
MAX_RETRIES      = 3


# ── Provider detection ─────────────────────────────────────────────────────────

def detect_provider(model: str) -> str:
    m = model.lower()
    if m.startswith("gpt") or m.startswith("o1") or m.startswith("o3"):
        return "openai"
    if m.startswith("gemini"):
        return "gemini"
    if "llama" in m or "mixtral" in m or "gemma" in m or "qwen" in m:
        return "groq"
    raise ValueError(f"Cannot detect provider for model '{model}'. "
                     f"Pass --provider openai|gemini|groq explicitly.")


def build_client(provider: str, model: str):
    if provider == "openai":
        from openai import OpenAI
        key = os.getenv("OPENAI_API_KEY")
        if not key:
            print("Error: OPENAI_API_KEY not set in .env"); sys.exit(1)
        return OpenAI(api_key=key)

    if provider == "gemini":
        from google import genai
        key = os.getenv("GEMINI_API_KEY")
        if not key:
            print("Error: GEMINI_API_KEY not set in .env"); sys.exit(1)
        return genai.Client(api_key=key)

    if provider == "groq":
        from groq import Groq
        key = os.getenv("GROQ_API_KEY")
        if not key:
            print("Error: GROQ_API_KEY not set in .env"); sys.exit(1)
        return Groq(api_key=key)

    raise ValueError(f"Unknown provider: {provider}")


# ── API calls ──────────────────────────────────────────────────────────────────

def call_openai(client, model: str, batch: list) -> str:
    user_msg = json.dumps(batch, ensure_ascii=False)
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_msg},
        ],
        temperature=0,
        max_tokens=800,
    )
    return resp.choices[0].message.content.strip()


def call_gemini(client, model: str, batch: list) -> str:
    from google.genai import types
    user_msg = json.dumps(batch, ensure_ascii=False)
    # API requires full model path (models/...) if not already prefixed
    full_model = model if model.startswith("models/") else f"models/{model}"
    resp = client.models.generate_content(
        model=full_model,
        contents=user_msg,
        config=types.GenerateContentConfig(
            system_instruction=SYSTEM_PROMPT,
            temperature=0,
            max_output_tokens=800,
        ),
    )
    return resp.text.strip()


def call_groq(client, model: str, batch: list) -> str:
    user_msg = json.dumps(batch, ensure_ascii=False)
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_msg},
        ],
        temperature=0,
        max_tokens=800,
    )
    return resp.choices[0].message.content.strip()


def call_model(client, provider: str, model: str, batch: list) -> str:
    if provider == "openai":
        return call_openai(client, model, batch)
    if provider == "gemini":
        return call_gemini(client, model, batch)
    if provider == "groq":
        return call_groq(client, model, batch)


# ── Response parsing ───────────────────────────────────────────────────────────

def parse_response(raw: str, expected_ids: list) -> dict:
    clean = re.sub(r"^```[a-z]*\n?", "", raw.strip(), flags=re.MULTILINE)
    clean = re.sub(r"\n?```$", "", clean.strip())
    data  = json.loads(clean)
    result = {}
    for item in data:
        sid    = str(item.get("id", ""))
        labels = item.get("labels", ["None"])
        result[sid] = labels
    for sid in expected_ids:
        if sid not in result:
            result[sid] = ["None"]
    return result


def label_batch(client, provider: str, model: str, batch: list) -> dict:
    expected_ids = [b["id"] for b in batch]
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            raw = call_model(client, provider, model, batch)
            return parse_response(raw, expected_ids)
        except Exception as exc:
            if attempt < MAX_RETRIES:
                time.sleep(3 * attempt)
            else:
                print(f"  Batch failed after {MAX_RETRIES} attempts: {exc}")
                return {sid: ["None"] for sid in expected_ids}


# ── Gold set loading ───────────────────────────────────────────────────────────

def _to_binary(v) -> int:
    if v is None:
        return 0
    s = str(v).strip()
    if s in ("", "None", "nan", "0"):
        return 0
    try:
        return 1 if float(s) >= 0.5 else 0
    except (ValueError, TypeError):
        return 0


def load_gold() -> pd.DataFrame:
    wb = openpyxl.load_workbook(GOLD_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rows.append(row)
    df = pd.DataFrame(rows)
    for col in FRAME_COLS.values():
        if col in df.columns:
            df[col] = df[col].apply(_to_binary)
    keep = ["sentence_id", "sentence_text"] + [c for c in FRAME_COLS.values() if c in df.columns]
    return df[keep]


# ── Metrics ────────────────────────────────────────────────────────────────────

def compute_metrics(gold: pd.DataFrame, predictions: dict) -> list:
    rows = []
    print(f"\n{'Label':<25} {'Precision':>9} {'Recall':>7} {'F1':>7} {'Accuracy':>9} "
          f"{'Human+':>7} {'LLM+':>6}")
    print("─" * 75)

    for frame, col in FRAME_COLS.items():
        if col not in gold.columns:
            continue
        y_true, y_pred = [], []
        for _, row in gold.iterrows():
            sid    = str(row["sentence_id"])
            labels = predictions.get(sid, ["None"])
            y_true.append(int(row[col]))
            y_pred.append(1 if frame in labels else 0)

        n        = len(y_true)
        accuracy = sum(t == p for t, p in zip(y_true, y_pred)) / n
        prec     = precision_score(y_true, y_pred, zero_division=0)
        rec      = recall_score(y_true, y_pred, zero_division=0)
        f1       = f1_score(y_true, y_pred, zero_division=0)
        h_pos    = sum(y_true)
        l_pos    = sum(y_pred)

        print(f"{frame:<25} {prec:>9.3f} {rec:>7.3f} {f1:>7.3f} {accuracy:>9.3f} "
              f"{h_pos:>7} {l_pos:>6}")
        rows.append({
            "label": frame, "precision": round(prec, 4), "recall": round(rec, 4),
            "f1": round(f1, 4), "accuracy": round(accuracy, 4),
            "human_positive": h_pos, "llm_positive": l_pos, "n": n,
        })

    macro_p  = sum(r["precision"] for r in rows) / len(rows)
    macro_r  = sum(r["recall"]    for r in rows) / len(rows)
    macro_f1 = sum(r["f1"]        for r in rows) / len(rows)
    print("─" * 75)
    print(f"{'MACRO AVG':<25} {macro_p:>9.3f} {macro_r:>7.3f} {macro_f1:>7.3f}")
    return rows


def print_comparison(alt_rows: list, model: str):
    if not HAIKU_CSV.exists():
        print("\n(Haiku baseline not found — skipping comparison)")
        return
    haiku = pd.read_csv(HAIKU_CSV).set_index("label")
    alt   = pd.DataFrame(alt_rows).set_index("label")

    print(f"\n── Side-by-side: Haiku vs {model} ──")
    print(f"{'Label':<25} {'Haiku P':>8} {'Alt P':>6}  {'Haiku R':>8} {'Alt R':>6}  "
          f"{'Haiku F1':>9} {'Alt F1':>7}")
    print("─" * 80)
    for frame in FRAMES:
        if frame not in haiku.index or frame not in alt.index:
            continue
        h, a = haiku.loc[frame], alt.loc[frame]
        print(f"{frame:<25} {h['precision']:>8.3f} {a['precision']:>6.3f}  "
              f"{h['recall']:>8.3f} {a['recall']:>6.3f}  "
              f"{h['f1']:>9.3f} {a['f1']:>7.3f}")
    h_macro = haiku[["precision", "recall", "f1"]].mean()
    a_macro = alt[["precision", "recall", "f1"]].mean()
    print("─" * 80)
    print(f"{'MACRO AVG':<25} {h_macro['precision']:>8.3f} {a_macro['precision']:>6.3f}  "
          f"{h_macro['recall']:>8.3f} {a_macro['recall']:>6.3f}  "
          f"{h_macro['f1']:>9.3f} {a_macro['f1']:>7.3f}")

    both_conservative = a_macro["precision"] > a_macro["recall"]
    print(f"\nRobustness check: {model} also conservative (precision > recall)? "
          f"{'YES ✓' if both_conservative else 'NO ✗'}")


# ── Main ───────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--model",    required=True, help="Model ID (e.g. gpt-4o-mini, gemini-2.0-flash, llama-3.3-70b-versatile)")
    p.add_argument("--provider", default=None,  help="openai|gemini|groq (auto-detected if omitted)")
    p.add_argument("--limit",    type=int, default=None, help="Use only first N gold sentences (test)")
    return p.parse_args()


def main():
    args     = parse_args()
    model    = args.model
    provider = args.provider or detect_provider(model)

    print(f"Provider: {provider}  |  Model: {model}")
    client = build_client(provider, model)

    print(f"Loading gold set ({GOLD_PATH.name}) …")
    gold = load_gold()
    if args.limit:
        gold = gold.head(args.limit)
    print(f"  {len(gold)} sentences")

    print(f"\nLabeling with {model} …")
    predictions = {}
    for i in range(0, len(gold), BATCH_SIZE):
        chunk = gold.iloc[i : i + BATCH_SIZE]
        batch = [{"id": str(r["sentence_id"]), "text": str(r["sentence_text"])}
                 for _, r in chunk.iterrows()]
        result = label_batch(client, provider, model, batch)
        predictions.update(result)
        done = min(i + BATCH_SIZE, len(gold))
        print(f"  {done}/{len(gold)} sentences labeled …", flush=True)
        time.sleep(INTER_CALL_DELAY)

    print(f"\n── {model} results vs human gold ──")
    rows = compute_metrics(gold, predictions)
    print_comparison(rows, model)

    safe_name = model.replace("/", "-").replace(":", "-")
    out_path  = OUTPUTS_TABLES / f"llm_validation_{safe_name}.csv"
    OUTPUTS_TABLES.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(out_path, index=False)
    print(f"\nSaved → {out_path}")


if __name__ == "__main__":
    main()
