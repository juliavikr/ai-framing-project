"""
clean_corpus.py — Build a clean sentence pool for annotation.

STAGE 1: Reads corpus.csv, splits to sentences, applies 5 quality filters,
         saves data/annotation/clean_sentences.csv.

STAGE 2: Redraws the 100-sentence Kappa overlap set from the clean pool
         (34 commercial / 33 policy / 33 public) and writes
         data/annotation/kappa_overlap_person_a_v3.xlsx and _b_v3.xlsx.

Usage:
    python src/processing/clean_corpus.py
    python src/processing/clean_corpus.py --no-langdetect   # faster, skips lang filter
    python src/processing/clean_corpus.py --stage1-only
"""

import argparse
import re
import sys
import uuid
from pathlib import Path

import nltk
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from src.config import CORPUS_CSV, DATA_ANNOTATION

for _r in ("punkt", "punkt_tab"):
    try:
        nltk.data.find(f"tokenizers/{_r}")
    except (LookupError, OSError):
        nltk.download(_r, quiet=True)

CLEAN_SENTENCES_CSV = DATA_ANNOTATION / "clean_sentences.csv"
MIN_WORDS = 8

# ── Filter patterns ────────────────────────────────────────────────────────────

_JUNK_RE = re.compile(
    r'(\{"|\bclassName\b|"div"|"span"|\s=>\s|/\*|https?://|©|@md:|px-|col-span)',
    re.IGNORECASE,
)

_AI_RE = re.compile(
    r'\b(AI|artificial intelligence|machine learning|model|media|'
    r'Microsoft|Meta\b|Google|DeepMind|OpenAI|Anthropic|Nvidia|'
    r'ChatGPT|AGI|superintelligence|neural|algorithm|automation|'
    r'robot|compute|LLM|GPT|Claude|Gemini|Llama|language model|'
    r'data science|deep learning|generative)\b',
    re.IGNORECASE,
)

_OFFTOPIC_RE = re.compile(
    r'(Public Assistance|emergency protective measures|Category [A-Z]\b|'
    r'Federal funding is available|trade deficit|GDP grew|unemployment rate)',
    re.IGNORECASE,
)

# Strong framing markers — absence → sentence more likely to be None
_FRAME_STRONG_RE = re.compile(
    r'\b(will|must|should|revolutionize|transform|trillion|billion|percent|'
    r'risk|harm|bias|discriminat|superintelligence|AGI|align|dangerous|'
    r'regulation|governance|oversight|existential|breakthrough|advance)\b',
    re.IGNORECASE,
)


def _is_header_fragment(text: str, wc: int) -> bool:
    if wc >= 6:
        return False
    letters = [c for c in text if c.isalpha()]
    if not letters:
        return True
    return sum(1 for c in letters if c.isupper()) / len(letters) > 0.7


# ── Stage 1 — filtering ────────────────────────────────────────────────────────

def apply_filters(text: str, use_langdetect: bool) -> str:
    """Return the name of the first failing filter, or '' if sentence passes all."""
    wc = len(text.split())

    if wc < MIN_WORDS:
        return "too_short"

    if _JUNK_RE.search(text):
        return "junk_fragment"

    if _OFFTOPIC_RE.search(text) or _is_header_fragment(text, wc):
        return "off_topic"

    if not _AI_RE.search(text):
        return "no_ai_relevance"

    if use_langdetect:
        try:
            from langdetect import detect_langs
            from langdetect import DetectorFactory
            DetectorFactory.seed = 0
            langs = detect_langs(text)
            top = langs[0]
            if top.lang != "en" or top.prob < 0.9:
                return "non_english"
        except Exception:
            pass  # detection failure → keep sentence

    return ""


def is_likely_none(text: str, wc: int) -> bool:
    """Heuristic: short sentence with no strong framing keywords → probably None."""
    return wc <= 20 and not bool(_FRAME_STRONG_RE.search(text))


def build_clean_pool(use_langdetect: bool) -> pd.DataFrame:
    print(f"\nReading corpus from {CORPUS_CSV} …")
    corpus = pd.read_csv(CORPUS_CSV, dtype=str)
    print(f"  {len(corpus):,} documents")

    print("Tokenising sentences …")
    rows = []
    for _, doc in corpus.iterrows():
        text = str(doc.get("text", "")) if pd.notna(doc.get("text", "")) else ""
        for sent in nltk.sent_tokenize(text):
            sent = sent.strip()
            rows.append({
                "sentence_id":  str(uuid.uuid4()),
                "doc_id":       doc["doc_id"],
                "actor":        doc["actor"],
                "actor_type":   doc["actor_type"],
                "context":      doc["context"],
                "date":         doc["date"],
                "post_chatgpt": doc["post_chatgpt"],
                "sentence_text": sent,
                "word_count":   len(sent.split()),
            })
    total = len(rows)
    print(f"  {total:,} sentences before filtering")

    print("Applying filters …" + (" (langdetect active — may take a few minutes)" if use_langdetect else " (langdetect disabled)"))
    filter_counts = {
        "too_short":       0,
        "junk_fragment":   0,
        "off_topic":       0,
        "no_ai_relevance": 0,
        "non_english":     0,
    }
    clean = []
    for i, r in enumerate(rows):
        if i % 20000 == 0 and i > 0:
            print(f"    … {i:,} / {total:,}")
        failed = apply_filters(r["sentence_text"], use_langdetect)
        if failed:
            filter_counts[failed] += 1
        else:
            r["likely_none"] = is_likely_none(r["sentence_text"], r["word_count"])
            clean.append(r)

    df = pd.DataFrame(clean)

    removed = total - len(df)
    print(f"\n── Filter summary ──────────────────────────────")
    print(f"  Total sentences       : {total:>8,}")
    for name, n in filter_counts.items():
        pct = n / total * 100
        print(f"  {name:<20}  {n:>7,}  ({pct:.1f}%)")
    print(f"  {'─' * 40}")
    print(f"  Removed (total)       : {removed:>7,}  ({removed/total*100:.1f}%)")
    print(f"  Kept (clean pool)     : {len(df):>7,}  ({len(df)/total*100:.1f}%)")
    print(f"  Likely None flagged   : {df['likely_none'].sum():>7,}")

    DATA_ANNOTATION.mkdir(parents=True, exist_ok=True)
    df.to_csv(CLEAN_SENTENCES_CSV, index=False)
    print(f"\nSaved → {CLEAN_SENTENCES_CSV}")
    return df


# ── Stage 2 — stratified overlap redraw ───────────────────────────────────────

CONTEXT_COUNTS = {"commercial": 34, "policy": 33, "public": 33}
MIN_ACTORS     = 5
MIN_NONE_FRAC  = 0.15

LABEL_COLS = ["Innovation_Progress", "Economic_Benefit", "Risk_Harm",
              "Regulation_Governance", "Existential_AGI", "None", "Notes"]
KEEP_COLS  = ["sentence_id", "actor", "context", "date", "sentence_text"]
ALL_COLS   = KEEP_COLS + LABEL_COLS

COL_WIDTHS = {
    "sentence_id": 36, "actor": 18, "context": 12, "date": 12,
    "sentence_text": 70, "Innovation_Progress": 10, "Economic_Benefit": 10,
    "Risk_Harm": 10, "Regulation_Governance": 14, "Existential_AGI": 10,
    "None": 7, "Notes": 30,
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
LABEL_FILL  = PatternFill("solid", fgColor="2E75B6")
NONE_FILL   = PatternFill("solid", fgColor="7F7F7F")
NOTES_FILL  = PatternFill("solid", fgColor="404040")
CTX_FILLS   = {
    "commercial": PatternFill("solid", fgColor="E2EFDA"),
    "policy":     PatternFill("solid", fgColor="FCE4D6"),
    "public":     PatternFill("solid", fgColor="EDE7F6"),
}
ROW_EVEN = PatternFill("solid", fgColor="EBF3FB")
ROW_ODD  = PatternFill("solid", fgColor="FFFFFF")
THIN     = Border(
    left=Side(style="thin", color="CCCCCC"),  right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),   bottom=Side(style="thin", color="CCCCCC"),
)

INSTRUCTIONS = [
    ["AI Framing Study — Kappa Overlap v3 (clean pool)"],
    [],
    ["IMPORTANT — READ BEFORE ANNOTATING"],
    ["This is round 3 of the Kappa calibration. Sentences were drawn from a"],
    ["quality-filtered pool — garbage, non-English, and off-topic sentences removed."],
    ["Re-read annotation_guidelines.md before starting, especially the None Boundary"],
    ["rules and the 10 Calibration Examples at the bottom."],
    ["Annotate independently — do NOT discuss with your partner before submitting."],
    [],
    ["HOW TO ANNOTATE"],
    ["1. Read each sentence in column E (sentence_text)."],
    ["2. Enter 1 in every applicable frame column; leave blank if not applicable."],
    ["3. A sentence can have multiple frames simultaneously."],
    ["4. If NO frame applies, enter 1 in 'None'. About 20-30% should be None."],
    ["5. Use Notes for genuinely ambiguous cases — flag them, don't guess."],
    [],
    ["FRAME DEFINITIONS"],
    ["Innovation_Progress",   "Explicit claim: AI advances society / science"],
    ["Economic_Benefit",      "Explicit claim: jobs, revenue, productivity, competitiveness"],
    ["Risk_Harm",             "Concrete near-term harm: bias, job loss, misuse, surveillance"],
    ["Regulation_Governance", "Explicit institutional mechanism: law, oversight, standards"],
    ["Existential_AGI",       "Civilizational-scale risk, AGI, superintelligence"],
    [],
    ["CONTEXT COLOURS: Green = commercial | Orange = policy | Purple = public"],
]


def sample_context_block(pool: pd.DataFrame, n: int, rng: int) -> pd.DataFrame:
    """Sample n sentences with actor diversity and likely-None quota from pool."""
    n_none   = max(int(np.ceil(n * MIN_NONE_FRAC)), 1)
    n_normal = n - n_none

    none_pool   = pool[pool["likely_none"] == True]
    normal_pool = pool[pool["likely_none"] == False]

    if len(none_pool) < n_none:
        n_none   = len(none_pool)
        n_normal = n - n_none

    sampled_none = none_pool.sample(n=n_none, random_state=rng, replace=False)

    avail_normal = normal_pool[~normal_pool["sentence_id"].isin(sampled_none["sentence_id"])]
    if len(avail_normal) < n_normal:
        avail_normal = pool[~pool["sentence_id"].isin(sampled_none["sentence_id"])]

    sampled_normal = avail_normal.sample(n=n_normal, random_state=rng, replace=False)
    result = pd.concat([sampled_none, sampled_normal], ignore_index=True)

    # Retry up to 15 times if actor diversity not met
    for attempt in range(1, 16):
        if result["actor"].nunique() >= MIN_ACTORS:
            break
        result = pool.sample(n=n, random_state=rng + attempt, replace=False)

    return result


def redraw_overlap(df_clean: pd.DataFrame, rng: int = 42) -> pd.DataFrame:
    parts = []
    for ctx, n in CONTEXT_COUNTS.items():
        pool  = df_clean[df_clean["context"] == ctx].copy()
        block = sample_context_block(pool, n, rng)
        parts.append(block)
    sampled = pd.concat(parts, ignore_index=True)
    sampled = sampled.sample(frac=1, random_state=rng).reset_index(drop=True)
    return sampled


def write_xlsx(df: pd.DataFrame, out_path: Path, tab_label: str, annotator: str) -> None:
    src = df[KEEP_COLS].copy()
    for col in LABEL_COLS:
        src[col] = ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_label

    for ci, col in enumerate(ALL_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN
        if col == "Notes":   cell.fill = NOTES_FILL
        elif col == "None":  cell.fill = NONE_FILL
        elif col in LABEL_COLS: cell.fill = LABEL_FILL
        else:                cell.fill = HEADER_FILL
    ws.row_dimensions[1].height = 32

    for ri, (_, row) in enumerate(src.iterrows(), 2):
        ctx  = str(row.get("context", ""))
        base = CTX_FILLS.get(ctx, ROW_EVEN if ri % 2 == 0 else ROW_ODD)
        for ci, col in enumerate(ALL_COLS, 1):
            val  = row[col] if col in row.index else ""
            cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
            cell.border = THIN
            if col in LABEL_COLS:
                cell.fill      = ROW_EVEN if ri % 2 == 0 else ROW_ODD
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(size=11)
            else:
                cell.fill      = base
                cell.alignment = Alignment(vertical="top", wrap_text=(col == "sentence_text"))
                cell.font      = Font(size=9)
        ws.row_dimensions[ri].height = 45

    for ci, col in enumerate(ALL_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)
    ws.freeze_panes = "F2"

    info = wb.create_sheet("Instructions")
    for r, row_data in enumerate(INSTRUCTIONS, 1):
        for c, val in enumerate(row_data, 1):
            cell = info.cell(row=r, column=c, value=val)
            if r in (1, 3, 10, 17):
                cell.font = Font(bold=True, size=11 if r == 1 else 10)
    info.cell(row=len(INSTRUCTIONS) + 2, column=1,
              value=f"Annotator: {annotator}").font = Font(bold=True)
    info.column_dimensions["A"].width = 42
    info.column_dimensions["B"].width = 55

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def print_overlap_summary(df: pd.DataFrame) -> None:
    print(f"\n── Overlap set summary ─────────────────────────")
    print(f"  Total sentences : {len(df)}")

    print(f"\n  Context breakdown:")
    for ctx in ("commercial", "policy", "public"):
        n = (df["context"] == ctx).sum()
        likely_none = df[df["context"] == ctx]["likely_none"].sum() if "likely_none" in df.columns else "—"
        print(f"    {ctx:<12}  {n:>3}  (likely-None: {likely_none})")

    print(f"\n  Actor breakdown:")
    actor_ctx = df.groupby(["actor", "context"]).size().unstack(fill_value=0)
    for actor, row in actor_ctx.iterrows():
        parts = "  ".join(f"{c}={v}" for c, v in row.items() if v > 0)
        print(f"    {actor:<20} {parts}")

    if "post_chatgpt" in df.columns:
        post = df["post_chatgpt"].astype(str).eq("1").sum()
        pre  = len(df) - post
        print(f"\n  post-ChatGPT: {post}  |  pre-ChatGPT: {pre}")


# ── Main ───────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Build clean sentence pool and redraw Kappa overlap set.")
    p.add_argument("--no-langdetect", action="store_true",
                   help="Skip language detection (faster, may keep a few non-English sentences)")
    p.add_argument("--stage1-only", action="store_true",
                   help="Run Stage 1 (filtering) only, skip Stage 2 (redraw)")
    p.add_argument("--seed", type=int, default=42, help="Random seed (default: 42)")
    return p.parse_args()


def main():
    args = parse_args()
    use_langdetect = not args.no_langdetect

    # ── Stage 1 ──────────────────────────────────────────────────────────────
    print("═" * 52)
    print("STAGE 1 — Build clean sentence pool")
    print("═" * 52)

    if CLEAN_SENTENCES_CSV.exists():
        print(f"\nClean pool already exists: {CLEAN_SENTENCES_CSV}")
        ans = input("Re-run filtering? (y/N) ").strip().lower()
        if ans != "y":
            print("Loading existing clean pool …")
            df_clean = pd.read_csv(CLEAN_SENTENCES_CSV, dtype=str)
            df_clean["likely_none"] = df_clean["likely_none"].map(
                {"True": True, "False": False, True: True, False: False}
            )
            print(f"  {len(df_clean):,} sentences loaded")
        else:
            df_clean = build_clean_pool(use_langdetect)
    else:
        df_clean = build_clean_pool(use_langdetect)

    if args.stage1_only:
        return

    # ── Stage 2 ──────────────────────────────────────────────────────────────
    print("\n" + "═" * 52)
    print("STAGE 2 — Redraw Kappa overlap set (v3)")
    print("═" * 52)

    overlap = redraw_overlap(df_clean, rng=args.seed)
    print_overlap_summary(overlap)

    out_a = DATA_ANNOTATION / "kappa_overlap_person_a_v3.xlsx"
    out_b = DATA_ANNOTATION / "kappa_overlap_person_b_v3.xlsx"

    write_xlsx(overlap, out_a, "Round 3 — Person A", "A (Julia)")
    write_xlsx(overlap, out_b, "Round 3 — Person B", "B (Elisabeth)")

    print(f"\n── Files saved ─────────────────────────────────")
    print(f"  {out_a}")
    print(f"  {out_b}")
    print(f"\n  Both files contain the same 100 sentences.")
    print(f"  All label columns are empty — annotate independently.\n")


if __name__ == "__main__":
    main()
